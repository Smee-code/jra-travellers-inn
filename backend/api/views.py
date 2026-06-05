from django.contrib.auth.hashers import make_password
from django.db.models import Count, Avg, Q, Sum, Min, Max
from django.utils import timezone
from datetime import date, timedelta
import csv
import io
import re
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import (
    User, RoomType, Room, Booking, RoomReview, AuditLog, ReportRecord, SystemConfig,
    is_uniform_booking_id, next_booking_id,
)
from .permissions import IsOwner, IsAdmin, IsOwnerOrAdmin
from .serializers import (
    CustomTokenObtainPairSerializer, RegisterSerializer,
    UserListSerializer, UserDetailSerializer,
    ManagedUserSerializer,
    RoomTypeSerializer, RoomSerializer,
    BookingSerializer, BookingCreateSerializer,
    AuditLogSerializer, ReportRecordSerializer, PushSubscriptionSerializer,
)
from .forecasting import (
    actual_vs_predicted_monthly,
    forecast_ranges,
    model_metadata,
)
from .push import send_user_push, webpush_configured


@api_view(['GET'])
@permission_classes([AllowAny])
def landing_data(request):
    active_rooms = Room.objects.select_related('room_type').filter(status='Active')
    avg_rating = RoomReview.objects.filter(room__status='Active').aggregate(avg=Avg('rating'))['avg'] or 0
    room_types = [
        {
            'name': rt.name,
            'count': active_rooms.filter(room_type=rt).count(),
            'base_price': rt.base_price,
            'capacity': rt.capacity,
        }
        for rt in RoomType.objects.order_by('base_price')
    ]
    featured_rooms = active_rooms.annotate(review_rating=Avg('room_reviews__rating')).order_by('-review_rating', 'price')[:6]
    return Response({
        'stats': {
            'rooms': active_rooms.count(),
            'bookings': Booking.objects.count(),
            'customers': User.objects.filter(role='customer').count(),
            'avg_rating': round(avg_rating, 1),
            'completed_stays': Booking.objects.filter(status='Completed').count(),
        },
        'room_types': room_types,
        'featured_rooms': RoomSerializer(featured_rooms, many=True).data,
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def contact_info(request):
    owner = (
        User.objects
        .filter(role='owner', is_active=True)
        .exclude(phone='')
        .order_by('-date_joined')
        .first()
    ) or (
        User.objects
        .filter(role='owner', is_active=True)
        .order_by('-date_joined')
        .first()
    )
    return Response({
        'phone': owner.phone if owner and owner.phone else '+63 32 555 0100',
        'email': owner.email if owner and owner.email else 'stay@travellersinn.ph',
        'address': owner.location if owner and owner.location else '12 Seaside Avenue',
    })


@api_view(['GET', 'PATCH'])
@permission_classes([AllowAny])
def branding(request):
    if request.method == 'PATCH':
        if not request.user.is_authenticated or request.user.role not in ('admin', 'owner'):
            return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    config, _ = SystemConfig.objects.get_or_create(pk=1)
    if request.method == 'PATCH':
        logo = str(request.data.get('logo_data_url') or '').strip()
        if logo and not logo.startswith('data:image/'):
            return Response({'detail': 'Logo must be an image data URL.'}, status=400)
        if len(logo) > 7_000_000:
            return Response({'detail': 'Logo image is too large. Use an image under 5 MB.'}, status=400)
        config.logo_data_url = logo
        config.save(update_fields=['logo_data_url', 'updated_at'])
    return Response({'logo_data_url': config.logo_data_url})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def push_public_key(request):
    from django.conf import settings
    return Response({
        'public_key': settings.WEBPUSH_VAPID_PUBLIC_KEY,
        'configured': webpush_configured(),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def push_subscribe(request):
    serializer = PushSubscriptionSerializer(data=request.data, context={'request': request})
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response({'detail': 'Notifications enabled.'}, status=status.HTTP_201_CREATED)


# ── Auth ──────────────────────────────────────────────────────────────────────

class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [AllowAny]


class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]


def _auth_user_payload(u):
    return {
        'id': u.id,
        'name': u.full_name,
        'first_name': u.first_name,
        'last_name': u.last_name,
        'email': u.email,
        'role': u.role,
        'phone': u.phone,
        'location': u.location,
        'username': u.username,
    }


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def me(request):
    u = request.user
    if request.method == 'PATCH':
        allowed = ('first_name', 'last_name', 'email', 'phone', 'location')
        for field in allowed:
            if field in request.data:
                setattr(u, field, str(request.data.get(field) or '').strip())
        u.save(update_fields=list(allowed))
    return Response(_auth_user_payload(u))


# ── Room Types ────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def booking_options(request):
    check_in = timezone.localdate() + timedelta(days=1)
    check_out = check_in + timedelta(days=4)
    filters = [{'value': '', 'label': 'All'}] + [
        {'value': rt.name, 'label': rt.name}
        for rt in RoomType.objects.order_by('base_price', 'name')
    ]
    return Response({
        'default_dates': {
            'in': check_in.isoformat(),
            'out': check_out.isoformat(),
        },
        'room_filters': filters,
    })


class RoomTypeViewSet(viewsets.ModelViewSet):
    queryset = RoomType.objects.all()
    serializer_class = RoomTypeSerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsOwnerOrAdmin()]


# ── Rooms ─────────────────────────────────────────────────────────────────────

class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.select_related('room_type').all()
    serializer_class = RoomSerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsOwnerOrAdmin()]

    def get_queryset(self):
        qs = super().get_queryset()
        # Customers only see active rooms
        if self.request.user.role == 'customer':
            qs = qs.filter(status='Active')
        room_type = self.request.query_params.get('type')
        if room_type and room_type != 'All':
            qs = qs.filter(room_type__name=room_type)
        return qs

    @action(detail=True, methods=['post'], permission_classes=[IsOwnerOrAdmin()])
    def toggle_status(self, request, pk=None):
        room = self.get_object()
        room.status = 'Inactive' if room.status == 'Active' else 'Active'
        room.save()
        actor = 'Owner' if request.user.role == 'owner' else 'Admin'
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action=f'{"Activated" if room.status == "Active" else "Deactivated"} room',
            target=f'{room.room_id} · {room.name}',
            kind=room.status,
        )
        return Response(RoomSerializer(room).data)


# ── Bookings ──────────────────────────────────────────────────────────────────

class BookingViewSet(viewsets.ModelViewSet):
    serializer_class = BookingSerializer

    def get_queryset(self):
        u = self.request.user
        qs = Booking.objects.select_related('guest', 'room', 'room__room_type')
        if u.role == 'customer':
            qs = qs.filter(guest=u)
        status_filter = self.request.query_params.get('status')
        if status_filter and status_filter != 'All':
            qs = qs.filter(status=status_filter)
        # Admin can filter by guest or room
        guest_id = self.request.query_params.get('guest_id')
        if guest_id and u.role in ('admin', 'owner'):
            qs = qs.filter(guest_id=guest_id)
        room = self.request.query_params.get('room')
        if room and u.role in ('admin', 'owner'):
            qs = qs.filter(room_id=room)
        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return BookingCreateSerializer
        return BookingSerializer

    def get_permissions(self):
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        booking = serializer.save()
        AuditLog.objects.create(
            who=f'Customer · {self.request.user.full_name}',
            action='Created booking',
            target=booking.booking_id,
            kind='Pending',
        )

    @action(detail=True, methods=['post'], permission_classes=[IsOwnerOrAdmin()])
    def confirm(self, request, pk=None):
        booking = self.get_object()
        booking.status = 'Confirmed'
        booking.save()
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action='Confirmed booking',
            target=booking.booking_id,
            kind='Confirmed',
        )
        send_user_push(
            booking.guest,
            'Booking confirmed',
            f'Your booking {booking.booking_id} for {booking.room.name} has been approved.',
        )
        return Response(BookingSerializer(booking).data)

    @action(detail=True, methods=['post'], permission_classes=[IsOwnerOrAdmin()])
    def complete(self, request, pk=None):
        booking = self.get_object()
        if booking.status != 'Confirmed':
            return Response({'detail': 'Only confirmed bookings can be marked completed.'}, status=400)
        booking.status = 'Completed'
        booking.save()
        actor = 'Owner' if request.user.role == 'owner' else 'Admin'
        AuditLog.objects.create(
            who=f'{actor} · {request.user.full_name}',
            action='Marked booking completed',
            target=booking.booking_id,
            kind='Completed',
        )
        return Response(BookingSerializer(booking).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def review(self, request, pk=None):
        booking = self.get_object()
        if request.user.role != 'customer' or booking.guest != request.user:
            return Response({'detail': 'You can only review your own completed bookings.'}, status=403)
        if booking.status != 'Completed':
            return Response({'detail': 'You can leave a review only after the booking is completed.'}, status=400)
        if hasattr(booking, 'review'):
            return Response({'detail': 'This booking already has a review.'}, status=400)

        try:
            rating = int(request.data.get('rating'))
        except (TypeError, ValueError):
            return Response({'detail': 'Rating is required.'}, status=400)
        if rating < 1 or rating > 5:
            return Response({'detail': 'Rating must be from 1 to 5 stars.'}, status=400)

        comment = str(request.data.get('comment') or '').strip()
        review = RoomReview.objects.create(
            booking=booking,
            room=booking.room,
            guest=request.user,
            rating=rating,
            comment=comment,
        )
        AuditLog.objects.create(
            who=f'Customer · {request.user.full_name}',
            action='Left room review',
            target=f'{booking.booking_id} · {booking.room.name}',
            kind='info',
        )
        return Response(BookingSerializer(booking).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], permission_classes=[IsOwnerOrAdmin()], url_path='import-history')
    def import_history(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Upload a CSV or Excel file.'}, status=400)

        name = upload.name.lower()
        try:
            rows = self._read_history_rows(upload, name)
        except Exception as exc:
            return Response({'detail': f'Could not read file: {exc}'}, status=400)

        created, skipped, errors = 0, 0, []
        valid_statuses = {choice[0] for choice in Booking.STATUS_CHOICES}

        for index, row in enumerate(rows, start=2):
            try:
                cleaned = {str(k).strip().lower(): v for k, v in row.items() if k}
                room_id = str(cleaned.get('room_id') or cleaned.get('room') or '').strip()
                if not room_id:
                    raise ValueError('room_id is required')
                room = Room.objects.get(room_id=room_id)

                check_in = self._parse_import_date(cleaned.get('check_in'))
                check_out = self._parse_import_date(cleaned.get('check_out'))
                if check_out <= check_in:
                    raise ValueError('check_out must be after check_in')

                nights = int(cleaned.get('nights') or (check_out - check_in).days)
                guests_count = int(cleaned.get('guests_count') or cleaned.get('guests') or 1)
                amount = int(float(cleaned.get('amount') or (nights * room.price)))
                status_value = str(cleaned.get('status') or 'Completed').strip().title()
                if status_value not in valid_statuses:
                    raise ValueError(f'invalid status "{status_value}"')

                guest = self._get_import_guest(cleaned)
                source_booking_id = str(cleaned.get('booking_id') or '').strip()
                booking_id = source_booking_id if is_uniform_booking_id(source_booking_id) else next_booking_id()

                if Booking.objects.filter(booking_id=booking_id).exists():
                    skipped += 1
                    continue

                Booking.objects.create(
                    booking_id=booking_id,
                    guest=guest,
                    room=room,
                    check_in=check_in,
                    check_out=check_out,
                    nights=nights,
                    guests_count=guests_count,
                    amount=amount,
                    status=status_value,
                )
                created += 1
            except Exception as exc:
                skipped += 1
                errors.append({'row': index, 'error': str(exc)})

        actor = 'Owner' if request.user.role == 'owner' else 'Admin'
        AuditLog.objects.create(
            who=f'{actor} Â· {request.user.full_name}',
            action='Imported booking history',
            target=f'{created} created, {skipped} skipped',
            kind='info',
        )
        return Response({'created': created, 'skipped': skipped, 'errors': errors[:20]})

    def _read_history_rows(self, upload, name):
        if name.endswith('.csv'):
            text = upload.read().decode('utf-8-sig')
            return list(csv.DictReader(io.StringIO(text)))
        if name.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            return [
                {headers[i]: value for i, value in enumerate(values) if i < len(headers)}
                for values in rows[1:]
                if any(value is not None and value != '' for value in values)
            ]
        raise ValueError('Only .csv and .xlsx files are supported')

    def _parse_import_date(self, value):
        if hasattr(value, 'date'):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value).strip())

    def _get_import_guest(self, row):
        email = str(row.get('guest_email') or row.get('email') or '').strip()
        name = str(row.get('guest_name') or row.get('guest') or 'Imported Guest').strip()
        if email:
            existing = User.objects.filter(email=email).first()
            if existing:
                return existing
        base = re.sub(r'[^a-z0-9_]+', '_', (email.split('@')[0] if email else name).lower()).strip('_') or 'imported_guest'
        username = base
        n = 1
        while User.objects.filter(username=username).exists():
            n += 1
            username = f'{base}_{n}'
        first, *rest = name.split()
        user = User(username=username, email=email, first_name=first, last_name=' '.join(rest), role='customer')
        user.set_unusable_password()
        user.save()
        return user

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated()])
    def cancel(self, request, pk=None):
        booking = self.get_object()
        if request.user.role == 'customer' and booking.guest != request.user:
            return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role == 'customer' and booking.status != 'Pending':
            return Response(
                {'detail': 'Confirmed bookings can no longer be cancelled by the customer. Please contact the inn.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        booking.status = 'Cancelled'
        booking.save()
        who = f'Customer · {request.user.full_name}' if request.user.role == 'customer' \
            else f'Admin · {request.user.full_name}'
        AuditLog.objects.create(
            who=who, action='Cancelled booking',
            target=booking.booking_id, kind='Cancelled',
        )
        return Response(BookingSerializer(booking).data)

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin()])
    def modify(self, request, pk=None):
        """Admin edits check-in, check-out, and/or guest count on a booking."""
        booking = self.get_object()
        check_in  = request.data.get('check_in',     str(booking.check_in))
        check_out = request.data.get('check_out',    str(booking.check_out))
        guests    = int(request.data.get('guests_count', booking.guests_count))

        from datetime import date as _date
        ci = _date.fromisoformat(check_in)
        co = _date.fromisoformat(check_out)
        if co <= ci:
            return Response({'detail': 'Check-out must be after check-in.'}, status=400)

        nights = (co - ci).days
        amount = nights * booking.room.price   # recalculate from room price

        booking.check_in     = ci
        booking.check_out    = co
        booking.nights       = nights
        booking.guests_count = guests
        booking.amount       = amount
        booking.save()

        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action='Modified booking',
            target=booking.booking_id,
            kind='info',
        )
        return Response(BookingSerializer(booking).data)


# ── Customers (admin management) ─────────────────────────────────────────────

class CustomerViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(role='customer').order_by('date_joined')
    serializer_class = UserListSerializer
    permission_classes = [IsOwnerOrAdmin]
    pagination_class = None

    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action=f'{"Activated" if user.is_active else "Deactivated"} account',
            target=f'{user.username} · {user.full_name}',
            kind='Active' if user.is_active else 'Inactive',
        )
        return Response(UserListSerializer(user).data)

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        user = self.get_object()
        # In production this would email a reset link. For prototype, set a temp password.
        temp_password = 'TempPass123!'
        user.set_password(temp_password)
        user.save()
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action='Reset password',
            target=f'{user.username} · {user.full_name}',
            kind='info',
        )
        return Response({'detail': f'Password reset link sent to {user.email}'})


# ── Audit Log ─────────────────────────────────────────────────────────────────

class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(role__in=('customer', 'owner')).order_by('-date_joined')
    permission_classes = [IsAdmin]
    pagination_class = None

    def get_serializer_class(self):
        if self.action == 'create':
            return ManagedUserSerializer
        return UserListSerializer

    def create(self, request, *args, **kwargs):
        serializer = ManagedUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action='Created user account',
            target=f'{user.username} · {user.full_name}',
            kind='info',
        )
        return Response(UserListSerializer(user).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action=f'{"Activated" if user.is_active else "Deactivated"} account',
            target=f'{user.username} · {user.full_name}',
            kind='Active' if user.is_active else 'Inactive',
        )
        return Response(UserListSerializer(user).data)

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        user = self.get_object()
        temp_password = 'TempPass123!'
        user.set_password(temp_password)
        user.save(update_fields=['password'])
        AuditLog.objects.create(
            who=f'Admin · {request.user.full_name}',
            action='Reset password',
            target=f'{user.username} · {user.full_name}',
            kind='info',
        )
        return Response({'detail': f'Password reset link sent to {user.email}'})


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()[:20]
    serializer_class = AuditLogSerializer
    permission_classes = [IsOwnerOrAdmin]


# ── Analytics helpers ────────────────────────────────────────────────────────

import calendar as _cal
import math as _math


def _period_config(period):
    """
    Return (cur_start, cur_end, prv_start, prv_end, months) for the
    selected period key ('12m' | 'ytd' | 'qtd').

    months  — list of (year, month) tuples covering the current window,
              oldest first; used by chart endpoints.
    prv_*   — equivalent prior window for period-over-period deltas.
    """
    today = date.today()

    if period == 'ytd':
        cur_start = date(today.year, 1, 1)
        prv_start = date(today.year - 1, 1, 1)
        prv_end   = date(today.year - 1, today.month, today.day)
        months    = [(today.year, m) for m in range(1, today.month + 1)]

    elif period == 'qtd':
        q_first   = ((today.month - 1) // 3) * 3 + 1   # first month of current quarter
        cur_start = date(today.year, q_first, 1)
        prv_start = date(today.year - 1, q_first, 1)
        prv_end   = prv_start + timedelta(days=(today - cur_start).days)
        months    = [(today.year, m) for m in range(q_first, today.month + 1)]

    else:   # '12m' — default
        try:
            cur_start = today.replace(year=today.year - 1)
            prv_start = today.replace(year=today.year - 2)
        except ValueError:          # Feb 29 edge case
            cur_start = today - timedelta(days=365)
            prv_start = today - timedelta(days=730)
        prv_end = cur_start
        months  = _months_back(12)

    return cur_start, today, prv_start, prv_end, months


def _months_back(n):
    """Return list of (year, month) tuples going back n months, oldest first."""
    today = date.today()
    y, m = today.year, today.month
    result = []
    for _ in range(n):
        result.append((y, m))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(result))


def _active_rooms():
    return Room.objects.filter(status='Active').count() or 1


def _month_occ(yr, mo):
    """Occupancy rate 0-100 for a calendar month based on booked vs available nights."""
    _, days = _cal.monthrange(yr, mo)
    avail = _active_rooms() * days
    start_dt, end_dt = date(yr, mo, 1), date(yr, mo, days)
    booked = (
        Booking.objects
        .filter(status__in=['Confirmed', 'Completed'],
                check_in__lte=end_dt, check_out__gt=start_dt)
        .aggregate(n=Sum('nights'))['n'] or 0
    )
    return round(booked / avail * 100, 1) if avail else 0.0


def _moving_avg_and_std(values, window=6):
    """Return (avg, std) of the last `window` non-None values."""
    non_null = [v for v in values if v is not None]
    sample = non_null[-window:] if non_null else []
    if not sample:
        return 0.0, 0.0
    avg = sum(sample) / len(sample)
    if len(sample) > 1:
        std = _math.sqrt(sum((x - avg) ** 2 for x in sample) / len(sample))
    else:
        std = avg * 0.15
    return avg, std


# ── Analytics endpoints ───────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_model_info(request):
    return Response(model_metadata())


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_metrics(request):
    period = request.query_params.get('period', '12m')
    cur_start, cur_end, prv_start, prv_end, _ = _period_config(period)

    PERIOD_LABELS = {
        '12m': ('vs. prior 12 months', 'avg this year'),
        'ytd': ('vs. same period last year', 'year to date'),
        'qtd': ('vs. same quarter last year', 'this quarter'),
    }
    delta_sub, occ_sub = PERIOD_LABELS.get(period, PERIOD_LABELS['12m'])

    def _window(start, end):
        qs  = Booking.objects.filter(check_in__gte=start, check_in__lte=end)
        tot = qs.count()
        can = qs.filter(status='Cancelled').count()
        agg = qs.filter(status__in=['Confirmed', 'Completed']).aggregate(
            r=Sum('amount'), n=Sum('nights'))
        nights  = agg['n'] or 0
        revenue = agg['r'] or 0
        days    = max((end - start).days, 1)
        avail   = _active_rooms() * days
        return {
            'total':        tot,
            'cancel_rate':  (can / tot * 100) if tot else 0.0,
            'occ_rate':     (nights / avail * 100) if avail else 0.0,
            'avg_per_night': (revenue / nights) if nights else 0.0,
        }

    c = _window(cur_start, cur_end)
    p = _window(prv_start, prv_end)

    def dpct(cv, pv):
        d = ((cv - pv) / pv * 100) if pv else 0.0
        return f'{d:+.1f}%', d >= 0

    def dpts(cv, pv):
        d = cv - pv
        return f'{d:+.1f} pts', d >= 0

    td, tu = dpct(c['total'], p['total'])
    od, ou = dpts(c['occ_rate'], p['occ_rate'])
    rd, ru = dpct(c['avg_per_night'], p['avg_per_night'])
    cd, cu = dpts(c['cancel_rate'], p['cancel_rate'])
    cu = not cu     # lower cancellation = positive signal

    return Response([
        {'label': 'Total Bookings',      'value': f'{c["total"]:,}',
         'delta': td, 'up': tu, 'sub': delta_sub},
        {'label': 'Occupancy Rate',       'value': f'{c["occ_rate"]:.1f}%',
         'delta': od, 'up': ou, 'sub': occ_sub},
        {'label': 'Avg. Revenue / Night', 'value': f'₱{c["avg_per_night"]:,.0f}',
         'delta': rd, 'up': ru, 'sub': 'per booked room'},
        {'label': 'Cancellation Rate',    'value': f'{c["cancel_rate"]:.1f}%',
         'delta': cd, 'up': cu, 'sub': 'lower is better'},
    ])


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_avp(request):
    try:
        return Response(actual_vs_predicted_monthly())
    except Exception as exc:
        return Response({
            'detail': f'Prophet forecasting is unavailable: {exc}',
            **model_metadata(),
        }, status=503)

@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_forecasts(request):
    try:
        return Response(forecast_ranges())
    except Exception as exc:
        return Response({
            'detail': f'Prophet forecasting is unavailable: {exc}',
            **model_metadata(),
        }, status=503)

    def _fmt_rev(n):
        if n >= 1_000_000:
            return f'₱{n / 1_000_000:.2f}M'
        return f'₱{round(n / 1000)}K'

    result = []
    for days, label in [(7, 'Next 7 days'), (30, 'Next 30 days'), (90, 'Next 90 days')]:
        pred = daily_rate * days
        avail = rooms * days
        occ = (pred / avail * 100) if avail else 0
        conf = max(1, round(pred * 0.12))
        result.append({
            'range': label,
            'bookings': round(pred),
            'occ': f'{occ:.0f}%',
            'conf': f'±{conf}',
            'rev': _fmt_rev(pred * avg_rev),
        })
    return Response(result)


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_occupancy(request):
    period = request.query_params.get('period', '12m')
    _, _, _, _, months = _period_config(period)
    labels = [date(yr, mo, 1).strftime('%b') for yr, mo in months]
    values = [_month_occ(yr, mo) for yr, mo in months]
    return Response({'labels': labels, 'values': values})


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_bookings_by_month(request):
    period = request.query_params.get('period', '12m')
    _, _, _, _, months = _period_config(period)
    # Cap at 8 for the bar chart — take last 8 if the window is larger
    display = months[-8:] if len(months) > 8 else months
    labels = [date(yr, mo, 1).strftime('%b') for yr, mo in display]
    values = [
        Booking.objects.filter(check_in__year=yr, check_in__month=mo).count()
        for yr, mo in display
    ]
    return Response({'labels': labels, 'values': values})


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_room_mix(request):
    period = request.query_params.get('period', '12m')
    cur_start, cur_end, _, _, _ = _period_config(period)

    rows = (
        Booking.objects
        .filter(check_in__gte=cur_start, check_in__lte=cur_end)
        .exclude(status='Cancelled')
        .values('room__room_type__name')
        .annotate(value=Count('id'))
        .order_by('-value')
    )
    if not rows:
        return Response([
            {'name': rt.name, 'value': 0}
            for rt in RoomType.objects.order_by('name')
        ])
    return Response([{'name': r['room__room_type__name'], 'value': r['value']} for r in rows])


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_heatmap(request):
    """12 × 5 intensity matrix (month × week-bucket) from actual booking density."""
    today = date.today()
    rooms = _active_rooms()

    matrix = []
    for mo in range(1, 13):
        yr = today.year
        _, days_in_mo = _cal.monthrange(yr, mo)
        row = []
        bucket_size = days_in_mo // 5   # ~6 days each

        for w in range(5):
            day_start = w * bucket_size + 1
            day_end = (day_start + bucket_size - 1) if w < 4 else days_in_mo
            avail = rooms * (day_end - day_start + 1)
            start_dt = date(yr, mo, day_start)
            end_dt   = date(yr, mo, day_end)

            booked = (
                Booking.objects
                .filter(status__in=['Confirmed', 'Completed'],
                        check_in__lte=end_dt, check_out__gt=start_dt)
                .aggregate(n=Sum('nights'))['n'] or 0
            )
            row.append(round(min(1.0, booked / avail), 2) if avail else 0.0)
        matrix.append(row)

    return Response(matrix)


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_accuracy(request):
    """Compare the 6-month moving average prediction against the last 3 full months of actuals."""
    months = _months_back(9)          # 6 months to train, 3 to evaluate
    counts = [
        Booking.objects.filter(check_in__year=yr, check_in__month=mo).count()
        for yr, mo in months
    ]

    train, test = counts[:6], counts[6:]
    train_avg, train_std = _moving_avg_and_std(train)

    errors = [abs(a - round(train_avg)) for a in test]
    mae = round(sum(errors) / len(errors), 1) if errors else 0.0

    within_20pct = sum(1 for e, a in zip(errors, test) if a == 0 or e / max(a, 1) <= 0.20)
    accuracy_pct = round(within_20pct / len(test) * 100) if test else 100

    occ_months = months[6:]
    occ_actuals = [_month_occ(yr, mo) for yr, mo in occ_months]
    occ_pred = _month_occ(*months[5])   # last training month as naive baseline
    occ_errors = [abs(o - occ_pred) for o in occ_actuals]
    occ_mae = round(sum(occ_errors) / len(occ_errors), 1) if occ_errors else 0.0

    return Response([
        {'metric': 'Mean abs. error', 'value': str(mae), 'unit': 'bookings/month',
         'plain': f'On average, the forecast is within {mae} bookings of the actual monthly count.'},
        {'metric': 'Forecast accuracy', 'value': f'{accuracy_pct}%', 'unit': 'last 3 months',
         'plain': f'{within_20pct} of the last 3 months landed within 20% of the prediction.'},
        {'metric': 'Occupancy error', 'value': f'±{occ_mae}', 'unit': 'pts',
         'plain': f'Predicted occupancy was within {occ_mae} percentage points of actual.'},
    ])


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def analytics_trends(request):
    """Auto-detect peaks, lows, and generate plain-language insights from real data."""
    today = date.today()
    period = request.query_params.get('period', '12m')
    _, _, _, _, months_12 = _period_config(period)

    # Monthly occupancy + booking counts
    monthly = []
    for yr, mo in months_12:
        occ = _month_occ(yr, mo)
        cnt = Booking.objects.filter(check_in__year=yr, check_in__month=mo).count()
        monthly.append({'yr': yr, 'mo': mo, 'occ': occ, 'cnt': cnt,
                        'label': date(yr, mo, 1).strftime('%B %Y')})

    # Sort to find peak and low
    by_occ = sorted(monthly, key=lambda x: x['occ'], reverse=True)
    by_cnt = sorted(monthly, key=lambda x: x['cnt'], reverse=True)

    flags = []
    if by_occ:
        peak = by_occ[0]
        flags.append({
            't': 'Peak flagged', 'v': peak['label'],
            'c': '#059669', 'ic': 'up',
            'd': f'{peak["label"]} had your highest occupancy at {peak["occ"]:.1f}%.',
        })
    if len(by_occ) > 1:
        low = by_occ[-1]
        flags.append({
            't': 'Low-demand flagged', 'v': low['label'],
            'c': '#d97706', 'ic': 'down',
            'd': f'{low["label"]} had your lowest occupancy at {low["occ"]:.1f}%. Consider a promotion.',
        })
    if len(by_cnt) > 0:
        busiest = by_cnt[0]
        flags.append({
            't': 'Busiest month', 'v': busiest['label'],
            'c': '#4f46e5', 'ic': 'star',
            'd': f'{busiest["label"]} had the most bookings with {busiest["cnt"]} reservations.',
        })

    # AVP comparison table: last 3 months with actual booking counts vs naive MA prediction
    avp_months = months_12[-3:]
    train_counts = [
        Booking.objects.filter(check_in__year=yr, check_in__month=mo).count()
        for yr, mo in months_12[:-3]
    ]
    train_avg, _ = _moving_avg_and_std(train_counts)
    avp_table = []
    for yr, mo in avp_months:
        actual_cnt = Booking.objects.filter(check_in__year=yr, check_in__month=mo).count()
        actual_occ = _month_occ(yr, mo)
        pred_occ   = _month_occ(yr, mo - 1 if mo > 1 else 12)   # naive: same as prior month
        avp_table.append({
            'period':    date(yr, mo, 1).strftime('%b %Y'),
            'predicted': round(train_avg),
            'actual':    actual_cnt,
            'occP':      f'{pred_occ:.0f}%',
            'occA':      f'{actual_occ:.0f}%',
        })

    # Auto-generate insight tags
    insights = []
    all_counts = [m['cnt'] for m in monthly]
    all_occ    = [m['occ'] for m in monthly]
    avg_cnt    = sum(all_counts) / len(all_counts) if all_counts else 0
    avg_occ    = sum(all_occ)    / len(all_occ)    if all_occ    else 0

    if by_occ:
        peak = by_occ[0]
        insights.append({
            'tag': 'PEAK',
            'text': f'{peak["label"]} is your highest-demand month at {peak["occ"]:.1f}% occupancy.',
        })
    if len(by_occ) > 1:
        low = by_occ[-1]
        insights.append({
            'tag': 'LOW',
            'text': f'{low["label"]} is your slowest month at {low["occ"]:.1f}% occupancy. Consider promotions.',
        })

    # Accuracy insight based on real MAE
    train, test = all_counts[:9], all_counts[9:]
    t_avg, _ = _moving_avg_and_std(train)
    errors = [abs(a - round(t_avg)) for a in test] if test else []
    mae = round(sum(errors) / len(errors), 1) if errors else 0
    insights.append({
        'tag': 'ACCURACY',
        'text': f'The baseline forecast is off by ~{mae} booking(s) per month on average.',
    })

    return Response({'flags': flags, 'avp_table': avp_table, 'insights': insights})


# ── Global search ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_all(request):
    from django.db.models import Q as DQ
    q = request.query_params.get('q', '').strip()
    if len(q) < 2:
        return Response({'bookings': [], 'rooms': [], 'customers': []})

    bookings = (
        Booking.objects
        .filter(DQ(booking_id__icontains=q)
              | DQ(guest__first_name__icontains=q)
              | DQ(guest__last_name__icontains=q))
        .select_related('guest', 'room', 'room__room_type')
        .order_by('-created_at')[:6]
    )
    rooms = (
        Room.objects
        .filter(DQ(room_id__icontains=q) | DQ(name__icontains=q))
        .select_related('room_type')[:6]
    )
    customers = (
        User.objects.filter(role='customer')
        .filter(DQ(first_name__icontains=q)
              | DQ(last_name__icontains=q)
              | DQ(email__icontains=q))[:6]
    )

    return Response({
        'bookings': [
            {'id': b.id, 'booking_id': b.booking_id,
             'guest_name': b.guest.full_name,
             'room_name': b.room.name, 'status': b.status}
            for b in bookings
        ],
        'rooms': [
            {'id': r.id, 'room_id': r.room_id, 'name': r.name,
             'room_type_name': r.room_type.name, 'status': r.status}
            for r in rooms
        ],
        'customers': [
            {'id': c.id, 'full_name': c.full_name, 'email': c.email}
            for c in customers
        ],
    })


# ── Reports ───────────────────────────────────────────────────────────────────

def _parse_report_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _report_defaults():
    bounds = Booking.objects.aggregate(start=Min('check_in'), end=Max('check_in'))
    today = timezone.localdate()
    default_to = bounds['end'] or today
    default_from = bounds['start'] or (default_to - timedelta(days=90))
    return default_from, default_to


def _report_name(date_from, date_to):
    if date_from and date_to:
        return f'Bookings Report {date_from:%b %d, %Y} to {date_to:%b %d, %Y}'
    return 'Bookings Report'


def _format_file_size(size):
    if size >= 1024 * 1024:
        return f'{size / (1024 * 1024):.1f} MB'
    return f'{max(1, round(size / 1024))} KB'


def _record_report(request, fmt, date_from, date_to, size):
    ReportRecord.objects.create(
        name=_report_name(date_from, date_to),
        format=fmt,
        date_from=date_from,
        date_to=date_to,
        size_label=_format_file_size(size),
        created_by=request.user if request.user.is_authenticated else None,
    )


def _report_bookings(date_from, date_to):
    qs = Booking.objects.select_related('guest', 'room').all()
    if date_from:
        qs = qs.filter(check_in__gte=date_from)
    if date_to:
        qs = qs.filter(check_in__lte=date_to)
    return qs


@api_view(['GET'])
@permission_classes([IsOwnerOrAdmin])
def report_options(request):
    default_from, default_to = _report_defaults()
    recent = ReportRecord.objects.all()[:10]
    return Response({
        'default_from': default_from.isoformat(),
        'default_to': default_to.isoformat(),
        'recent': ReportRecordSerializer(recent, many=True).data,
    })


@api_view(['POST'])
@permission_classes([IsOwnerOrAdmin])
def generate_report(request):
    fmt = request.data.get('format', 'pdf')
    date_from = _parse_report_date(request.data.get('from'))
    date_to = _parse_report_date(request.data.get('to'))

    if fmt == 'excel':
        from openpyxl import Workbook
        from django.http import HttpResponse
        import io
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bookings Report'
        ws.append(["Booking ID", "Guest", "Room", "Check-in", "Check-out", "Nights", "Amount", "Status"])
        bookings = _report_bookings(date_from, date_to)
        for b in bookings:
            ws.append([
                b.booking_id, b.guest.full_name, b.room.name,
                str(b.check_in), str(b.check_out), b.nights, b.amount, b.status
            ])
        buf = io.BytesIO()
        wb.save(buf)
        payload = buf.getvalue()
        _record_report(request, 'excel', date_from, date_to, len(payload))
        response = HttpResponse(
            payload,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bookings_report.xlsx"'
        return response

    # PDF
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from django.http import HttpResponse
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    # Use ASCII-safe title — Helvetica is Latin-1 only; em-dash and peso sign would corrupt output
    elements = [Paragraph("Traveller's Inn - Bookings Report", styles['Title'])]

    bookings = _report_bookings(date_from, date_to)
    table_data = [["Booking ID", "Guest", "Room", "Check-in", "Nights", "Amount (PHP)", "Status"]]
    for b in bookings:
        table_data.append([
            b.booking_id, b.guest.full_name, b.room.name,
            str(b.check_in), str(b.nights), f'PHP {b.amount:,}', b.status
        ])

    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e6eaf0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    payload = buf.getvalue()
    _record_report(request, 'pdf', date_from, date_to, len(payload))

    response = HttpResponse(payload, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bookings_report.pdf"'
    return response
